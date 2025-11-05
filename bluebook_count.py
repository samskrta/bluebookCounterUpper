#!/usr/bin/env python3
"""
Blue Book Counter Upper
-----------------------

Small utility to parse the first sheet (typically named 'ALL') of an Excel
export and count the number of unique quotes created by each technician.

Assumptions based on the report structure:
- Each quote begins on a row where column B contains the creator string in the
  form "TECH NAME (HH:MM AM|PM)".
- Rows belonging to the same quote leave column B empty.
- Column B's header can be something like "Created %sBy (At)"; this is ignored.

Usage:
  python bluebook_count.py --file data/BlueBook_Report_10_01_2025.xlsx

Optional:
  --sheet ALL             Sheet name (defaults to first sheet if not found)
  --csv out.csv           Write results to a CSV file in addition to stdout

This script has no external dependencies beyond openpyxl.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import subprocess
import sys
from collections import Counter, OrderedDict
from pathlib import Path

from openpyxl import load_workbook


TECH_ENTRY_REGEX = re.compile(
    r"^(?P<name>.+?)\s*\((?P<time>\d{1,2}:\d{2}\s?[AP]M)\)$",
    re.IGNORECASE,
)


def extract_technician_name(cell_value: object) -> str | None:
    """Return technician name if the given cell value marks a new quote.

    A valid creator cell matches "<name> (<time>)" where <time> is like
    5:07 PM or 05:07PM. Returns the <name> portion or None if not a match.
    """

    if cell_value is None:
        return None

    text = str(cell_value).strip()
    if not text:
        return None

    # Skip common header text
    header_like = text.lower().replace(" ", "")
    if header_like.startswith("created%sby(at)".replace(" ", "")):
        return None

    match = TECH_ENTRY_REGEX.match(text)
    if match:
        name = match.group("name").strip()
        # Normalize excessive inner whitespace
        name = re.sub(r"\s+", " ", name)
        return name
    return None


def count_quotes_per_technician(xlsx_path: Path, sheet_name: str | None) -> OrderedDict[str, int]:
    """Parse the workbook and count quotes per technician.

    If sheet_name is provided but not found, the first sheet is used.
    Returns an OrderedDict sorted by descending count then by name.
    """

    if not xlsx_path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ws = None
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    counts: Counter[str] = Counter()

    # Iterate all rows, examine column B (index 2)
    for row in ws.iter_rows(values_only=True):
        col_b_value = row[1] if len(row) > 1 else None
        name = extract_technician_name(col_b_value)
        if name:
            counts[name] += 1

    # Sort by descending count then ascending name for stable output
    sorted_items = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0].lower()))
    ordered = OrderedDict(sorted_items)
    return ordered


def print_summary(results: OrderedDict[str, int]) -> None:
    total = sum(results.values())
    print("Technician,Quotes")
    for tech, count in results.items():
        print(f"{tech},{count}")
    print(f"TOTAL,{total}")


def write_csv(results: OrderedDict[str, int], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Technician", "Quotes"])
        for tech, count in results.items():
            writer.writerow([tech, count])
        writer.writerow(["TOTAL", sum(results.values())])


def _load_ws_for_comments(xlsx_path: Path, sheet_name: str | None):
    """Load worksheet with comments/formatting available.

    We load with data_only=False and read_only=False so that cell comments are
    accessible for change detection.
    """

    wb = load_workbook(filename=str(xlsx_path), data_only=False, read_only=False)
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.worksheets[0]


def _find_header_columns(ws) -> dict[str, int]:
    """Return header mapping for creator, top-level Labor and detail Labor columns.

    Keys returned (0-based indexes):
    - 'creator': column containing 'Created'
    - 'top_labor': the high-level 'Labor' column in the summary header row
    - 'detail_labor': the 'Labor' column inside the line-item subtable row
    Fallbacks are B (1) for creator, D (3) for top_labor and D (3) for detail_labor.
    """

    creator_idx = None
    top_labor_idx = None
    detail_labor_idx = None

    # Scan the first few rows to detect both header rows
    for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
        if not row or not any(v is not None for v in row):
            continue
        texts = [str(v).strip() if v is not None else "" for v in row]
        lowers = [t.lower() for t in texts]
        lower_set = set(lowers)

        # Creator column
        if creator_idx is None:
            for idx, low in enumerate(lowers):
                if low.startswith("created"):
                    creator_idx = idx
                    break

        # Top header row contains Service Charge / Parts / Total
        if top_labor_idx is None and {"service charge", "parts", "total"} & lower_set and "labor" in lower_set:
            top_labor_idx = lowers.index("labor")

        # Detail header row contains Job Name / Part Number / Part Price / Labor / Total
        if detail_labor_idx is None and {"job name", "part number", "part price", "labor", "total"} <= lower_set:
            # If multiple occurrences of 'labor', choose the last one (often farther right)
            cand = [i for i, v in enumerate(lowers) if v == "labor"]
            if cand:
                detail_labor_idx = cand[-1]

        if creator_idx is not None and top_labor_idx is not None and detail_labor_idx is not None:
            break

    # Fallbacks
    if creator_idx is None:
        creator_idx = 1
    if top_labor_idx is None:
        top_labor_idx = 3
    if detail_labor_idx is None:
        detail_labor_idx = top_labor_idx

    return {"creator": creator_idx, "top_labor": top_labor_idx, "detail_labor": detail_labor_idx}


def _parse_currency(text: str | None) -> float | None:
    if not text:
        return None
    # Require the number to NOT be immediately preceded by a letter (avoids F12 -> 12)
    matches = re.findall(r"(?<![A-Za-z])\$?\s*([0-9]+(?:\.[0-9]{1,2})?)", str(text))
    if not matches:
        return None
    try:
        return float(matches[-1])
    except ValueError:
        return None


def _parse_labor_code_and_amount(text: str | None, code_price_map: dict[str, float] | None = None) -> tuple[str | None, float | None]:
    """Extract labor code like 'J13' and amount like 133.65 from strings.

    Returns (code, amount). If pattern not found, falls back to currency parse.
    """
    if not text:
        return (None, None)
    s = str(text)
    m = re.search(r"([A-Z]\d{1,2})\s*[-\u2013\u2014]\s*\$?\s*([0-9]+(?:\.[0-9]{1,2})?)", s)
    if m:
        try:
            return (m.group(1), float(m.group(2)))
        except ValueError:
            pass
    # Code-only case
    m2 = re.search(r"^\s*([A-Z]\d{1,2})\s*$", s)
    if m2:
        code = m2.group(1)
        amount = None
        if code_price_map and code in code_price_map:
            amount = float(code_price_map[code])
        return (code, amount)
    return (None, _parse_currency(s))


def build_labor_code_price_map(ws) -> dict[str, float]:
    """Build a map of labor code (e.g., J13) to its price by scanning the sheet.

    Sources:
    - Any cell text containing patterns like "J13 - $133.65"
    - Any cell comments containing the same patterns
    Last observed value wins (they should be consistent).
    """
    code_to_price: dict[str, float] = {}
    pair_pattern = re.compile(r"([A-Z]\d{1,2})\s*[-\u2013\u2014]\s*\$?\s*([0-9]+(?:\.[0-9]{1,2})?)")
    for row in ws.iter_rows(values_only=False):
        for c in row:
            if c is None:
                continue
            for source in (c.value, c.comment.text if c.comment is not None else None):
                if source is None:
                    continue
                s = str(source)
                for m in pair_pattern.finditer(s):
                    code = m.group(1)
                    try:
                        amount = float(m.group(2))
                    except ValueError:
                        continue
                    code_to_price[code] = amount
    return code_to_price


def _parse_comment_change(comment_text: str, cell_amount: float | None) -> tuple[str, float | None, float | None]:
    """Infer direction ('up'|'down'|'even'|'unknown') and before/after amounts."""

    text = comment_text or ""
    m = re.search(
        r"from\s*\$?\s*([0-9]+(?:\.[0-9]{1,2})?)\s*(?:to|→|->)\s*\$?\s*([0-9]+(?:\.[0-9]{1,2})?)",
        text,
        re.IGNORECASE,
    )
    before = after = None
    if m:
        before = float(m.group(1))
        after = float(m.group(2))
    else:
        nums = re.findall(r"\$?\s*([0-9]+(?:\.[0-9]{1,2})?)", text)
        if len(nums) >= 2:
            before = float(nums[0])
            after = float(nums[-1])
        elif len(nums) == 1 and cell_amount is not None:
            before = float(nums[0])
            after = cell_amount

    if before is None or after is None:
        return ("unknown", before, after)
    if after > before:
        return ("up", before, after)
    if after < before:
        return ("down", before, after)
    return ("even", before, after)


def collect_modifications(xlsx_path: Path, sheet_name: str | None) -> list[dict[str, object]]:
    """Collect per-quote flags for labor change and naive direction.

    Keys: row, technician, labor_value, labor_modified, labor_direction,
    labor_before, labor_after, labor_comment, part_modified.
    """

    ws = _load_ws_for_comments(xlsx_path, sheet_name)
    col_map = _find_header_columns(ws)
    code_price_map = build_labor_code_price_map(ws)
    top_labor_col = col_map["top_labor"]
    detail_labor_col = col_map["detail_labor"]
    creator_col = col_map["creator"]

    # Realize all rows for block traversal
    all_rows = list(ws.iter_rows(values_only=False))

    records: list[dict[str, object]] = []
    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        creator_cell = row[creator_col] if len(row) > creator_col else None
        tech = extract_technician_name(creator_cell.value if creator_cell is not None else None)
        if not tech:
            i += 1
            continue

        # Identify block [i, j) until next technician row
        j = i + 1
        while j < len(all_rows):
            next_row = all_rows[j]
            next_creator_cell = next_row[creator_col] if len(next_row) > creator_col else None
            if extract_technician_name(next_creator_cell.value if next_creator_cell is not None else None):
                break
            j += 1

        # Header labor cell (treated as the current/after value)
        labor_cell = row[top_labor_col] if len(row) > top_labor_col else None
        labor_value_str = str(labor_cell.value).strip() if labor_cell and labor_cell.value is not None else ""
        labor_code, labor_amount = _parse_labor_code_and_amount(labor_value_str, code_price_map)

        labor_comment_text = labor_cell.comment.text if (labor_cell and labor_cell.comment) else ""
        labor_modified = bool(labor_comment_text)

        # Collect labor codes/amounts from detail sub-rows for "natural" inference
        detail_pairs: list[tuple[str | None, float | None]] = []
        for r in all_rows[i + 1:j]:
            if len(r) <= detail_labor_col:
                continue
            c = r[detail_labor_col]
            val = str(c.value).strip() if c and c.value is not None else ""
            detail_pairs.append(_parse_labor_code_and_amount(val, code_price_map))

        detail_amounts = [amt for (_code, amt) in detail_pairs if isinstance(amt, (int, float))]
        natural_amt = max(detail_amounts) if detail_amounts else None

        # If comment text doesn't specify before/after, infer using block information
        direction, before_amt, after_amt = _parse_comment_change(labor_comment_text, labor_amount)
        if direction == "unknown":
            # Prefer the natural highest detail amount if available
            if natural_amt is not None and labor_amount is not None:
                before_amt = float(natural_amt)
                after_amt = float(labor_amount)
                if abs(after_amt - before_amt) < 0.005:
                    direction = "even"
                else:
                    direction = "up" if after_amt > before_amt else "down"

        # Consider any other commented cell in the header row as a part/price change marker
        part_modified = False
        for c in row:
            if c is None:
                continue
            if labor_cell is not None and c.coordinate == labor_cell.coordinate:
                continue
            if c.comment is not None:
                part_modified = True
                break

        # Compute changed flag and require a comment to count as a modification
        changed = False
        if labor_modified:
            if before_amt is not None and (after_amt if after_amt is not None else labor_amount) is not None:
                final_after = after_amt if after_amt is not None else labor_amount
                if final_after is not None and before_amt is not None:
                    changed = abs(float(final_after) - float(before_amt)) >= 0.005
        else:
            # No comment -> do not treat as modified; suppress direction to avoid false signal
            direction = "unknown"

        records.append(
            {
                "row": i + 1,
                "technician": tech,
                "labor_value": labor_value_str,
                "labor_code": labor_code or "",
                "labor_commented": labor_modified,
                "labor_changed": changed,
                "labor_direction": direction,
                "labor_before": before_amt,
                "labor_after": after_amt if after_amt is not None else labor_amount,
                "labor_comment": labor_comment_text,
                "part_modified": bool(part_modified),
            }
        )

        i = j

    return records


def write_modifications_csv(records: list[dict[str, object]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "row",
        "technician",
        "labor_value",
        "labor_commented",
        "labor_changed",
        "labor_direction",
        "labor_before",
        "labor_after",
        "part_modified",
        "labor_comment",
        "labor_code",
    ]
    with out_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in records:
            writer.writerow({k: rec.get(k, "") for k in fieldnames})


def summarize_modifications(records: list[dict[str, object]]) -> OrderedDict[str, dict[str, int]]:
    per_tech: dict[str, dict[str, int]] = {}
    for r in records:
        tech = str(r["technician"])
        per_tech.setdefault(tech, {"total": 0, "labor_modified": 0, "up": 0, "down": 0, "even": 0})
        per_tech[tech]["total"] += 1
        if r.get("labor_changed"):
            per_tech[tech]["labor_modified"] += 1
            direction = str(r.get("labor_direction") or "unknown")
            if direction in ("up", "down", "even"):
                per_tech[tech][direction] += 1
    ordered = OrderedDict(sorted(per_tech.items(), key=lambda kv: (-kv[1]["total"], kv[0].lower())))
    return ordered


def write_modifications_summary_csv(summary: OrderedDict[str, dict[str, int]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Technician", "Total", "Labor Modified", "% Modified", "Up", "Down", "Even"])
        for tech, d in summary.items():
            total = d["total"] or 1
            pct = round(100.0 * d["labor_modified"] / total, 2)
            writer.writerow([tech, d["total"], d["labor_modified"], pct, d["up"], d["down"], d["even"]])


def open_with_default_app(path: Path) -> None:
    """Best-effort open of the given path with the OS default app."""
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        elif os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception:
        # Non-fatal if we fail to open automatically
        pass


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Count Blue Book quotes per technician from Excel export.")
    default_file = Path("data/BlueBook_Report_10_01_2025.xlsx")
    parser.add_argument("--file", type=Path, default=None, help="Path to the Excel file (.xlsx)")
    parser.add_argument("--sheet", type=str, default="ALL", help="Sheet name to parse (default: ALL)")
    parser.add_argument("--csv", type=Path, default=None, help="Optional path to write counts CSV")
    parser.add_argument(
        "--mods-csv",
        type=Path,
        default=None,
        help="Optional path to write per-line modification flags CSV (_mods.csv by default)",
    )
    parser.add_argument(
        "--mods-summary-csv",
        type=Path,
        default=None,
        help="Optional path to write per-tech modification summary CSV (_mods_summary.csv by default)",
    )
    parser.add_argument("--gui", action="store_true", help="Use a simple file picker GUI and message box output")

    args = parser.parse_args(argv)

    selected_file: Path | None = args.file

    # GUI fallback: if --gui is set or --file is missing, prompt for a file via Tk
    if args.gui or selected_file is None:
        try:
            # Defer Tk imports so CLI environments without GUI don't pay the cost
            import tkinter as tk
            from tkinter import filedialog, messagebox

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            initial_dir = str(Path.cwd() / "data")
            selected_path = filedialog.askopenfilename(
                title="Select BlueBook Excel report",
                initialdir=initial_dir,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
            if not selected_path:
                messagebox.showwarning("BlueBook Counter Upper", "No file selected. Exiting.")
                return 1
            selected_file = Path(selected_path)
        except Exception as exc:
            print(f"GUI selection failed: {exc}", file=sys.stderr)
            # If GUI fails and no file was provided, abort
            if selected_file is None:
                return 1

    # Default CSV output path next to source file if not provided
    default_csv: Path | None = None
    if selected_file is not None:
        default_csv = selected_file.with_name(selected_file.stem + "_counts.csv")
        default_mods_csv = selected_file.with_name(selected_file.stem + "_mods.csv")
        default_mods_summary_csv = selected_file.with_name(selected_file.stem + "_mods_summary.csv")

    try:
        results = count_quotes_per_technician(selected_file, args.sheet)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    print_summary(results)

    out_csv_path: Path | None = args.csv or default_csv
    if out_csv_path is not None:
        try:
            write_csv(results, out_csv_path)
        except Exception as exc:
            print(f"Failed to write CSV: {exc}", file=sys.stderr)
            return 1
        # Best-effort auto-open of the generated CSV
        open_with_default_app(out_csv_path)

    # Collect and write modification details for labor/parts
    if selected_file is not None:
        try:
            records = collect_modifications(selected_file, args.sheet)
        except Exception as exc:
            print(f"Warning: failed to collect modifications: {exc}", file=sys.stderr)
            records = []

        mods_csv_path: Path | None = args.mods_csv or (default_mods_csv if selected_file is not None else None)
        if mods_csv_path is not None and records:
            try:
                write_modifications_csv(records, mods_csv_path)
                open_with_default_app(mods_csv_path)
            except Exception as exc:
                print(f"Failed to write modifications CSV: {exc}", file=sys.stderr)

        if records:
            mods_summary = summarize_modifications(records)
            mods_summary_path: Path | None = args.mods_summary_csv or default_mods_summary_csv
            if mods_summary_path is not None:
                try:
                    write_modifications_summary_csv(mods_summary, mods_summary_path)
                    open_with_default_app(mods_summary_path)
                except Exception as exc:
                    print(f"Failed to write modifications summary CSV: {exc}", file=sys.stderr)

    # If GUI mode, show a simple completion dialog
    if args.gui:
        try:
            import tkinter as tk
            from tkinter import messagebox

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            total = sum(results.values())
            message = (
                f"Finished. Found {total} quotes across {len(results)} technicians.\n\n"
                f"CSV saved to:\n{out_csv_path}"
            )
            messagebox.showinfo("BlueBook Counter Upper", message)
        except Exception:
            # Non-fatal if dialog fails
            pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())



